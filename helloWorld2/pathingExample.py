from pathlib import Path
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('tester2.xlsx')
charSheet = wb['Character']
historySheet = wb['Tester']

path = Path()
print(path.glob('*'))  # all files and directories
print(path.glob('*.*'))  # all files
print(path.glob('*.py'))  # all files of type

# print all the python files in the directory
for file in path.glob('*.py'):
    print(file)


def update_history(current_hp, x, new_hp):
    # add to end of sheet for history
    end_location = historySheet.max_row + 1
    historySheet.cell(end_location, 1).value = current_hp
    historySheet.cell(end_location, 2).value = x
    historySheet.cell(end_location, 3).value = new_hp


def change_hitpoints(x):
    current_hp = charSheet.cell(8, 1).value
    if x < 0:
        print('losing hit points!')
    new_hp = x + current_hp
    charSheet.cell(8, 1).value = new_hp
    # update history tab
    update_history(current_hp, x, new_hp)
    wb.save('tester2.xlsx')


print('losing 5 hitpoints!')
change_hitpoints(-5)
values = Reference(historySheet,
                   min_row=2,
                   max_row=historySheet.max_row,
                   min_col=3,
                   max_col=3
                   )

chart = BarChart()
chart.add_data(values)
historySheet.add_chart(chart, 5, 2)

"""
cell = sheet['a1']  # for some reason its backwards for the next part
cell2 = sheet.cell(2, 1)  # .cell is row THEN column unlike above!
print(cell.value)
print(cell2.value)
print('testing column...')
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    print(cell.value)
"""
