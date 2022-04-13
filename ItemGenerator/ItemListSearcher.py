from pathlib import Path
import openpyxl as xl


workbook_file_name = 'ItemLists.xlsx'
save_file_name = 'test1.xlsx'
# wondrous_sheet_name = 'WondrousItems'
# TODO workbook_file_path = ''


def process_workbook(wondrous_sheet_name, roll, prefix, affix, item_type):
    wb = xl.load_workbook(workbook_file_name)
    item_sheet = wb[wondrous_sheet_name]
    position_array = []
    item_name = "None"
    # print("process values for generator: " + str(roll) + ' ' + prefix + affix + item_type)  # debugging
    # find positions of acceptable items in spreadsheet

    for i in range(item_sheet.max_row):
        i_loc = i+1  # excel starts at 1 not 0

        if item_sheet.cell(i_loc, 1).value == item_type:
            if item_sheet.cell(i_loc, 2).value == prefix:
                if item_sheet.cell(i_loc, 3).value == affix:
                    position_array.append(i_loc)
    print(position_array)
    for j in range(len(position_array)):
        if roll <= item_sheet.cell(position_array[j], 4).value:
            item_name = item_sheet.cell(position_array[j], 5).value

    print(item_name)

    # chart = BarChart()
    # chart.add_data(values)
    # history_sheet.add_chart(chart, 'd4')
    # wb.save(workbook_file_name)

    return item_name
