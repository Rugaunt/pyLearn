from pathlib import Path
import openpyxl as xl


workbook_file_name = 'ItemLists.xlsx'
save_file_name = 'test1.xlsx'
# TODO workbook_file_path = ''



path = Path()
print(path.glob('*'))  # all files and directories
print(path.glob('*.*'))  # all files
print(path.glob('*.py'))  # all files of type

# print all the python files in the directory
for file in path.glob('*.py'):
    print(file)

"""
def process_workbook(filename):
    wb = xl.load_workbook(workbook_file_name)
    char_sheet = wb['Character']
    history_sheet = wb['Tester']
    print('losing 5 hitpoints!')
    change_hitpoints(-5, char_sheet, wb, history_sheet)
    values = Reference(history_sheet,
                       min_row=2,
                       max_row=history_sheet.max_row,
                       min_col=3,
                       max_col=3
                       )

    # chart = BarChart()
    chart.add_data(values)
    history_sheet.add_chart(chart, 'd4')
    wb.save(workbook_file_name)
"""

def update_history(current_hp, x, new_hp, history_sheet):
    # add to end of sheet for history
    end_location = history_sheet.max_row + 1
    history_sheet.cell(end_location, 1).value = current_hp
    history_sheet.cell(end_location, 2).value = x
    history_sheet.cell(end_location, 3).value = new_hp


def change_hitpoints(x, char_sheet, wb, history_sheet):
    current_hp = char_sheet.cell(8, 1).value
    if x < 0:
        print('losing hit points!')
    new_hp = x + current_hp
    char_sheet.cell(8, 1).value = new_hp
    # update history tab
    update_history(current_hp, x, new_hp, history_sheet)
    wb.save('tester2.xlsx')


process_workbook(workbook_file_name)


"""
cell = sheet['a1']  # for some reason its backwards for the next part
cell2 = sheet.cell(2, 1)  # .cell is row THEN column unlike above!
print(cell.value)
print(cell2.value)
print('testing column...')
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    print(cell.value)
"""
